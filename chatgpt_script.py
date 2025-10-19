"""
Filename: chatgpt_script.py
Author: Wesley Zheng
Date: 2025-4-20
Description: ​This script automates the process of submitting a prompt to ChatGPT, uploading a file, and pasting the response into a spreadsheet.
"""

import os
import pyautogui
import pyperclip
from openpyxl import load_workbook
import re

citizens = []
responses = []
pyautogui.FAILSAFE = False
pyautogui.PAUSE = 0.7
prompt = """Extract the following details from the FDA response document above: 1. Date of Response: Identify the date on which the FDA issued the response. 2. Responding FDA Center: Specify which FDA center or division responded to the petition (e.g., CDER, CBER, CDRH). 3. Response to Petition: Indicate the FDA's decision or action taken in response to the petition (e.g., approved, denied, partially approved, other). 4. Cited Statutes or Regulations: List all statutes or regulations cited by the FDA in its response. 5. Justification for Response: Summarize the reasoning or justifications provided by the FDA to support its decision. Ensure that the extracted information is accurate, relevant, and organized in a structured format, such as a bullet list or table. If any information is not explicitly stated, indicate 'Not Mentioned'. Remove markdown and put each numbered item in a Python list. Do not use nested lists. Do not include any other text such as comments or explanations. Put in code box."""

def get_file_names(folder_path):
    """
    Traverse the directory tree starting at 'folder_path' to identify and collect filenames matching specific patterns.

    - Appends filenames matching the pattern 'FDA-YYYY-P-NNNN-0001' (citizen petitions) to the 'citizens' list.
    - Tracks the most recent file matching 'from_FDA' and, upon encountering a file with a differing prefix, appends the previous 'from_FDA' file (ensuring that it is the final response from FDA) to the 'responses' list.
    - Ensures the last tracked 'from_FDA' file is added to 'responses' after traversal.

    Parameters:
    folder_path (str): The root directory path to begin the search.

    Note:
    This function relies on the global lists 'citizens' and 'responses' to store the results.
    """
    prev = None
    for _, _, files in os.walk(folder_path):
        prev = None
        for file in files:
            if re.search(r'FDA-\d{4}-P-\d{4}-0001', file) and int(re.findall(r'\d{4}', file)[1]) >= 4645:
                citizens.append(file)
            elif re.search(r'from_FDA', file):
                prev = file
            if prev and prev[:16] != file[:16]:
                if int(re.findall(r'\d{4}', file)[1]) >= 3545:
                    responses.append(prev)
                prev = None
    responses.append(prev)
        
            

def find_button(image_path, confidence = 0.7):
    """
    Locate a button on the screen by matching a provided image with a specified confidence level.

    Args:
    image_path (str): File path to the image of the button to locate.
    confidence (float, optional): Matching confidence threshold between 0 and 1. Defaults to 0.8.
    
    Returns:
    Box: A tuple (left, top, width, height) representing the coordinates of the located button.
    None: If the button is not found on the screen.
    
    Note:
    Requires OpenCV to be installed for the confidence parameter to work.
    If the image is not found, a message "Button not found." is printed.
    """
    button_location = pyautogui.locateOnScreen(image_path, confidence=confidence)
    if button_location is not None:
        return button_location
    else:
        print("Button not found.")

def move_mouse_to_button_and_click(image_path):
    """
    Locates a button on the screen using an image, moves the mouse to its center, and clicks it.

    Parameters:
        image_path (str): The file path to the image of the button to locate.

    Returns:
        None
    """
    button_location = find_button(image_path)
    if button_location is not None:
        x, y = pyautogui.center(button_location)
        pyautogui.moveTo(x, y, duration=0.5)
        pyautogui.click()
    else:
        print("No button location provided.")

def paste_text(text):
    """
    Simulates pasting the specified text into the active window by copying it to the clipboard and triggering the Ctrl+V hotkey.
    
    Args:
    text (str): The text to paste into the active window.
    """
    pyperclip.copy(text)
    pyautogui.hotkey('ctrl', 'v')

def detect_error(image_path, confidence = 0.7):
    """
    Checks if a specified image appears on the screen with a given confidence level.

    Args:
        image_path (str): Path to the image file to locate on the screen.
        confidence (float, optional): Matching confidence threshold (default is 0.7).

    Returns:
        bool: True if the image is found on the screen; False otherwise.
    """
    try:
        pyautogui.locateOnScreen(image_path, confidence=confidence)
        return True
    except Exception as e:
        return False

def wait_until_complete(image_path, confidence = 0.7):
    """
    Continuously monitors the screen for the appearance of a specific image, indicating task completion.
    
    Parameters:
        image_path (str): The file path of the image to detect on the screen.
        confidence (float, optional): The confidence threshold for image matching, between 0 and 1. Defaults to 0.95.
    
    Returns:
        bool: True if the image is found, indicating completion; False if an error image is detected.
    """
    while True:
        try:
            if detect_error("error.png"):
                return False
            button_location = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if button_location is not None:
                return True
        except Exception as e:
            continue


"""
Processes a list of citizen petition files and response files by automating interactions with a GUI to extract data 
from each file and save the results into corresponding Excel spreadsheets.

Steps:
1. Retrieves file names for citizen petitions and responses using `get_file_names`.
2. Iterates through the list of citizen petition files:
   - Simulates GUI interactions to upload the file and submit a prompt.
   - Waits for the process to complete or detects errors.
   - Extracts the response data, parses it, and saves it into the 'outputs_citizens.xlsx' spreadsheet.
3. Repeats the same process for response files, saving the results into the 'outputs_responses.xlsx' spreadsheet.

Parameters:
- folder (str): The folder containing the files to process.
- citizens (list): List of citizen petition file names.
- responses (list): List of response file names.

Note:
- Relies on GUI automation using `pyautogui` and clipboard operations with `pyperclip`.
- Requires Excel files ('outputs_citizens.xlsx' and 'outputs_responses.xlsx') with a sheet named '2024'.
"""
folder = "2019"
get_file_names(folder)
year = "2019"
print(len(responses), "files found")
counter = 133
total = len(responses)
for file_name in responses:
    move_mouse_to_button_and_click("prompt.png")
    paste_text(prompt)
    move_mouse_to_button_and_click("upload.png")
    move_mouse_to_button_and_click("upload_from_computer.png")
    import time
    time.sleep(1)
    paste_text(f"C:\\Users\\wesle\\OneDrive\\Desktop\\bclt\\{year}\\{file_name}")
    pyautogui.hotkey('enter')
    error = wait_until_complete("complete.png")
    if not error:
        move_mouse_to_button_and_click("unable_extract.png")
        pyautogui.moveTo(1100, 800, duration=0.5)
        pyautogui.click()
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.hotkey('backspace')
        continue
    move_mouse_to_button_and_click("ask.png")
    pyautogui.moveTo(200, 400, duration=0.5)
    wait_until_complete("voice.png")
    move_mouse_to_button_and_click("python_copy.png")
    lst = eval(pyperclip.paste())
    wb = load_workbook("outputs_responses.xlsx")
    ws = wb[year]
    ws.cell(row=counter, column=1).value = file_name
    for i, item in enumerate(lst):
        ws.cell(row=counter, column=i+2).value = item
    counter += 1
    wb.save("outputs_responses.xlsx")
    wb.close()

