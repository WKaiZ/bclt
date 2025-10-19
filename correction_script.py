"""
Filename: correction_script.py
Author: Wesley Zheng
Date: 2025-4-29
Description: ​This script automates the process of resubmitting a prompt to ChatGPT, uploading a file, and pasting the response into a spreadsheet.
"""

import pyautogui
import pyperclip
from openpyxl import load_workbook
import re


def read_lines_to_list(file_path):
    """
    Reads all lines from a text file and returns them as a list of strings, with trailing whitespace removed from each line.

    Parameters:
        file_path (str): The path to the text file to be read.

    Returns:
        list: A list of strings, where each string represents a line from the file with trailing whitespace removed.

    Example:
        lines = read_lines_to_list('example.txt')
        print(lines)  # Output: ['line1', 'line2', 'line3']
    """
    with open(file_path, 'r') as file:
        lines = [line.rstrip() for line in file]
    return lines

def find_button(image_path, confidence = 0.7):
    """
    See chatgpt_script.py.
    """
    button_location = pyautogui.locateOnScreen(image_path, confidence=confidence)
    if button_location is not None:
        return button_location
    else:
        print("Button not found.")

def move_mouse_to_button_and_click(image_path):
    """
    See chatgpt_script.py.
    """
    button_location = find_button(image_path)
    if button_location is not None:
        x, y = pyautogui.center(button_location)
        pyautogui.moveTo(x, y, duration=0.5)
        pyautogui.click()
    else:
        print("No button location provided.")

def paste_text(text):
    """
    See chatgpt_script.py.
    """
    pyperclip.copy(text)
    pyautogui.hotkey('ctrl', 'v')

def detect_error(image_path, confidence = 0.7):
    """
    See chatgpt_script.py.
    """
    try:
        pyautogui.locateOnScreen(image_path, confidence=confidence)
        return True
    except Exception as e:
        return False

def wait_until_complete(image_path, confidence = 0.7):
    """
    See chatgpt_script.py.
    """
    while True:
        try:
            if detect_error("error.png"):
                return False
            button_location = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if button_location is not None:
                return True
        except Exception as e:
            continue

def find_cell_coordinates(file_path, sheet_name, target_string):
    """
    Finds the coordinates of a cell in an Excel sheet that contains a specific target string.

    Parameters:
        file_path (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to search within.
        target_string (str): The string to search for in the sheet.

    Returns:
        tuple: A tuple (row, column) representing the coordinates of the cell containing the target string.
            Returns None if the target string is not found.

    Example:
        coordinates = find_cell_coordinates('example.xlsx', 'Sheet1', 'TargetValue')
        print(coordinates)  # Output: (3, 2) if the value is found in row 3, column 2.
    """
    wb = load_workbook(filename=file_path)
    sheet = wb[sheet_name]
    
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            if cell.value == target_string:
                return (cell.row, cell.column)
    return None

"""
Reads the contents of two text files and assigns their lines to the variables `citizens` and `responses`.

Files:
    'bad_summaries_petitions.txt': Contains a list of citizen petition file names.
    'bad_summaries_responses.txt': Contains a list of response file names.

Returns:
    citizens (list): A list of strings representing the lines from 'bad_summaries_petitions.txt'.
    responses (list): A list of strings representing the lines from 'bad_summaries_responses.txt'.

Example:
    citizens = ['petition1.txt', 'petition2.txt']
    responses = ['response1.txt', 'response2.txt']
"""
citizens = read_lines_to_list('bad_summaries_petitions.txt')
responses  = read_lines_to_list('bad_summaries_responses.txt')
pyautogui.FAILSAFE = False
pyautogui.PAUSE = 0.7
prompt = """Extract the following details from the FDA response document above: 1. Date of Response: Identify the date on which the FDA issued the response. 2. Responding FDA Center: Specify which FDA center or division responded to the petition (e.g., CDER, CBER, CDRH). 3. Response to Petition: Indicate the FDA's decision or action taken in response to the petition (e.g., approved, denied, partially approved, other). 4. Cited Statutes or Regulations: List all statutes or regulations cited by the FDA in its response. 5. Justification for Response: Summarize the reasoning or justifications provided by the FDA to support its decision. Ensure that the extracted information is accurate, relevant, and organized in a structured format, such as a bullet list or table. If any information is not explicitly stated, indicate 'Not Mentioned'. Remove markdown and put each numbered item in a Python list. Do not use nested lists. Do not include any other text such as comments or explanations. Put in code box."""

"""
See chatgpt_script.py.
"""
for file_name in responses:
    year = re.findall(r'\d{4}', file_name)[0]
    move_mouse_to_button_and_click("prompt.png")
    paste_text(prompt)
    move_mouse_to_button_and_click("upload.png")
    move_mouse_to_button_and_click("upload_from_computer.png")
    paste_text(f"C:\\Users\\wesle\\OneDrive\\Desktop\\bclt\\{year}\\{file_name}")
    pyautogui.hotkey('enter')
    error = wait_until_complete("complete.png")
    if not error:
        move_mouse_to_button_and_click("unable_extract.png")
        pyautogui.moveTo(1100, 800, duration=0.5)
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.hotkey('backspace')
        continue
    move_mouse_to_button_and_click("ask.png")
    wait_until_complete("voice.png")
    move_mouse_to_button_and_click("python_copy.png")
    lst = eval(pyperclip.paste())
    wb = load_workbook("outputs_responses.xlsx")
    ws = wb[year]
    r, _ = find_cell_coordinates("outputs_responses.xlsx", year, file_name)
    for i, item in enumerate(lst):
        if item == "Not Mentioned":
            print("Not Mentioned")
        ws.cell(row=r, column= i + 2).value = item
    wb.save("outputs_responses.xlsx")
    wb.close()
    print("Completed file: ", file_name)