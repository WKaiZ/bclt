import os
import pyautogui
import pyperclip
from openpyxl import load_workbook
import re

citizens = []
responses = []
pyautogui.FAILSAFE = False
pyautogui.PAUSE = 1
prompt = """Extract the following details from the FDA petition document above: 1. Date of Petition: Identify the date on which the petition was submitted. 2. Identity of Submitting Entity: Specify the name of the individual, company, or organization that submitted the petition. 3. Representation Details: Determine if the submitting entity is representing another entity (e.g., law firm representing a company). If so, provide the name of the represented entity. 4. Cited Statutes or Regulations: List all statutes or regulations cited by the petitioner to justify their request (e.g., 505(q), 21 C.F.R. 10.30). 5. FDA Action Commented On: Identify which action or policy of the FDA the petitioner is commenting on (e.g., notice to manufacturers, guidance for industry, regulation, establishment of a reference listed drug). 6. Requested Action: Specify the action the petitioner wants the FDA to take. 7. Justification for Request: Summarize the reasons or justifications provided by the petitioner for requesting this action. Remove markdown and put each numbered item in a Python list. Do not use nested lists. Do not include any other text such as comments or explanations."""

def get_file_names(folder_path):
    prev = None
    for _, _, files in os.walk(folder_path):
        for file in files:
            if re.search(r'FDA-2020-P-\d{4}-\d{4}_Citizen_Petition', file):
                citizens.append(file)
            elif re.search(r'from_FDA', file):
                prev = file
            if prev and prev[:16] != file[:16]:
                responses.append(prev)
                prev = None
    responses.append(prev)

def find_button(image_path, confidence = 0.8):
    button_location = pyautogui.locateOnScreen(image_path, confidence=confidence)
    if button_location is not None:
        return button_location
    else:
        print("Button not found.")

def move_mouse_to_button_and_click(image_path):
    button_location = find_button(image_path)
    if button_location is not None:
        x, y = pyautogui.center(button_location)
        pyautogui.moveTo(x, y, duration=0.5)
        pyautogui.click()
    else:
        print("No button location provided.")

def type_into_input_box(text, interval = 0.1):
    pyautogui.typewrite(text, interval)

def paste_text(text):
    pyperclip.copy(text)
    pyautogui.hotkey('ctrl', 'v')

def wait_until_complete(image_path, confidence = 0.95):
    while True:
        try:
            button_location = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if button_location is not None:
                return
        except Exception as e:
            continue

folder = "2020"
get_file_names(folder)
pyautogui.scroll(-5000)
counter = 2
for file_name in citizens:
    move_mouse_to_button_and_click("prompt.png")
    paste_text(prompt)
    move_mouse_to_button_and_click("upload.png")
    move_mouse_to_button_and_click("upload_from_computer.png")
    paste_text(file_name)
    move_mouse_to_button_and_click("open.png")
    wait_until_complete("complete.png")
    move_mouse_to_button_and_click("ask.png")
    wait_until_complete("voice.png")
    move_mouse_to_button_and_click("white.png")
    pyautogui.scroll(-5000)
    move_mouse_to_button_and_click("python_copy.png")
    lst = eval(pyperclip.paste())
    wb = load_workbook("outputs_citizens.xlsx")
    ws = wb.active
    ws.cell(row=counter, column=1).value = file_name
    for i, item in enumerate(lst):
        ws.cell(row=counter, column=i+2).value = item
    counter += 1
    wb.save("outputs_citizens.xlsx")
    wb.close()



