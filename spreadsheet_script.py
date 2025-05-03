"""
Filename: spreadsheet_script.py
Author: Wesley Zheng
Date: 2025-4-29
Description: ​This script automates the process of eliminating duplicate summaries for the same citizen petitions and responses.
"""

import re
from openpyxl import load_workbook

def clean_excel_cells(file_path, regex_pattern):
    """
    Cleans the content of cells in an Excel file by applying a regular expression pattern to remove unwanted text.

    Parameters:
        file_path (str): The path to the Excel file to be cleaned.
        regex_pattern (str): A regular expression pattern used to identify and remove unwanted text from cell values.

    Returns:
        None: The function modifies the Excel file in place and saves the changes.
    """
    wb = load_workbook(file_path)
    pattern = re.compile(regex_pattern)

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row = 2, min_col = 2):
            for cell in row:
                new_value = pattern.sub('', cell.value)
                if new_value != cell.value:
                    cell.value = new_value
    wb.save(file_path)

def collect_bad_summaries(file_path, regex_pattern):
    """
    Collects rows from an Excel file where at least one cell matches a given regular expression pattern.

    Parameters:
        file_path (str): The path to the Excel file to be processed.
        regex_pattern (str): A regular expression pattern used to identify matching cell values.

    Returns:
        list: A list of values from the first column of rows where at least one cell matches the pattern.

    Example:
        bad_summaries = collect_bad_summaries("example.xlsx", r'Not Mentioned')
        print(bad_summaries)  # Output: ['row1_value', 'row2_value', ...]
    """
    wb = load_workbook(file_path)
    pattern = re.compile(regex_pattern)

    files = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            if sum(pattern.search(cell) is not None for cell in row) >= 1:
                files.append(row[0])

    return files

def write_list_to_txt(data_list, file_path):
    """
    Writes a list of strings to a text file, with each item on a new line.

    Parameters:
        data_list (list): The list of strings to be written to the file.
        file_path (str): The path to the text file where the data will be written.

    Returns:
        None: The function writes the data to the file and does not return anything.

    Example:
        write_list_to_txt(['item1', 'item2', 'item3'], 'output.txt')
        # Creates a file 'output.txt' with each item on a new line.
    """
    with open(file_path, 'w', encoding='utf-8') as f:
        for item in data_list:
            f.write(f"{item}\n")



def remove_rows_with_duplicate_numbers(file_path):
    """
    Removes rows with duplicate numbers from an Excel file based on a specific column.

    Parameters:
        file_path (str): The path to the Excel file to be processed.

    Behavior:
        - If the file name contains 'citizens', duplicates are removed starting from the top of the sheet.
        - If the file name contains 'responses', duplicates are removed starting from the bottom of the sheet.
        - The function identifies duplicate numbers in the first column of each row using a four-digit pattern (e.g., a year).
        - Rows with duplicate numbers are deleted, keeping only the first occurrence (for 'citizens') or the last occurrence (for 'responses').

    Returns:
        None: The function modifies the Excel file in place and saves the changes.

    Example:
        remove_rows_with_duplicate_numbers("outputs_citizens.xlsx")
        # Removes rows with duplicate numbers in the 'outputs_citizens.xlsx' file.
    """
    wb = load_workbook(file_path)
    first_or_last = 1 if re.findall(r'citizens|responses', file_path)[0] == 'citizens' else -1
    for ws in wb.worksheets:
        seen_numbers = set()
        rows_to_delete = []
        if first_or_last == 1:
            for idx, row in enumerate(ws.iter_rows(min_row = 2), start = 2):
                row_has_duplicate = False
                row_numbers = set()

                number = re.findall(r'\d{4}', row[0].value)[1]
                if number in seen_numbers:
                    row_has_duplicate = True
                else:
                    row_numbers.add(number)

                if row_has_duplicate:
                    rows_to_delete.append(idx)
                else:
                    seen_numbers.update(row_numbers)
        else:
            for idx in range(ws.max_row, 1, -1):
                row = ws[idx]
                row_has_duplicate = False
                row_numbers = set()

                number = re.findall(r'\d{4}', row[0].value)[1]
                if number in seen_numbers:
                    row_has_duplicate = True
                else:
                    row_numbers.add(number)

                if row_has_duplicate:
                    rows_to_delete.append(idx)
                else:
                    seen_numbers.update(row_numbers)
                    
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

    wb.save(file_path)

"""
Performs a series of operations to clean and process Excel files, remove duplicate rows, and collect bad summaries.

Steps:
1. Cleans the content of cells in the Excel files "outputs_citizens.xlsx" and "outputs_responses.xlsx" by removing unwanted text using a regular expression pattern.
2. Removes rows with duplicate numbers from the Excel files:
   - For "outputs_citizens.xlsx", keeps the first occurrence of duplicate rows.
   - For "outputs_responses.xlsx", keeps the last occurrence of duplicate rows.
3. Collects rows from "outputs_citizens.xlsx" and "outputs_responses.xlsx" where at least one cell matches the pattern "Not Mentioned".
4. Writes the collected bad summaries to text files:
   - "bad_summaries_petitions.txt" for citizen petitions.
   - "bad_summaries_responses.txt" for responses.

Files:
    - "outputs_citizens.xlsx": Excel file containing citizen petition data.
    - "outputs_responses.xlsx": Excel file containing response data.
    - "bad_summaries_petitions.txt": Output text file for bad summaries of citizen petitions.
    - "bad_summaries_responses.txt": Output text file for bad summaries of responses.

"""
clean_excel_cells("outputs_citizens.xlsx", r'^.*:\s|^\d\.\s')
clean_excel_cells("outputs_responses.xlsx", r'^.*:\s|^\d\.\s')
remove_rows_with_duplicate_numbers("outputs_responses.xlsx")
remove_rows_with_duplicate_numbers("outputs_citizens.xlsx")
bad_summaries = collect_bad_summaries("outputs_citizens.xlsx", r'Not Mentioned')
write_list_to_txt(bad_summaries, "bad_summaries_petitions.txt")
bad_summaries = collect_bad_summaries("outputs_responses.xlsx", r'Not Mentioned')
write_list_to_txt(bad_summaries, "bad_summaries_responses.txt")

